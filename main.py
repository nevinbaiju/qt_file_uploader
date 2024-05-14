from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog

from openpyxl import load_workbook

import sys
import os
import json


class Set_Folder(QtWidgets.QWidget):
	def __init__(self):
		super().__init__()
		self.setupUi()
	def setupUi(self):
		self.setObjectName("MainWindow")
		self.resize(592, 368)
		
		self.title_label = QtWidgets.QLabel(self)
		self.title_label.setGeometry(QtCore.QRect(20, 20, 561, 71))
		font = QtGui.QFont()
		font.setPointSize(10)
		font.setBold(True)
		self.title_label.setFont(font)
		self.title_label.setObjectName("title_label")
		self.open_folder_button = QtWidgets.QPushButton(self)
		self.open_folder_button.setGeometry(QtCore.QRect(200, 150, 181, 41))
		self.open_folder_button.setObjectName("open_folder_button")
		self.continue_button = QtWidgets.QPushButton(self)
		self.continue_button.setGeometry(QtCore.QRect(400, 320, 181, 41))
		self.continue_button.setObjectName("open_folder_button")
		self.selected_folder_label = QtWidgets.QLabel(self)
		self.selected_folder_label.setGeometry(QtCore.QRect(120, 240, 361, 51))
		self.selected_folder_label.setObjectName("selected_folder_label")

		self.retranslateUi()
		QtCore.QMetaObject.connectSlotsByName(self)

	def retranslateUi(self):
		_translate = QtCore.QCoreApplication.translate
		self.setWindowTitle(_translate("MainWindow", "MainWindow"))
		self.title_label.setText(_translate("MainWindow", ""))
		self.open_folder_button.setText(_translate("MainWindow", "Open Folder"))
		self.continue_button.setText(_translate("MainWindow", "Continue"))
		self.selected_folder_label.setText(_translate("MainWindow", "Selected Folder:"))
		
		self.open_folder_button.clicked.connect(self.select_folder)
		
	def select_folder(self):
		folder_path = QFileDialog.getExistingDirectory(self, 'Open Folder', 'C:\\Users\\nevin\\Downloads')
		if folder_path:
			self.selected_folder_label.setText(f'Selected Folder: {folder_path}')
			with open('config.json', 'w') as file:
				json.dump({'output_folder': folder_path}, file)

class Upload_Files(QtWidgets.QWidget):
	def __init__(self):
		super().__init__()
		self.setupUi()
	def setupUi(self):
		self.setObjectName("MainWindow")
		self.resize(592, 368)
		
		self.title_label = QtWidgets.QLabel(self)
		self.title_label.setGeometry(QtCore.QRect(10, 20, 561, 31))
		font = QtGui.QFont()
		font.setPointSize(16)
		font.setBold(True)
		self.title_label.setFont(font)
		self.title_label.setObjectName("title_label")
		self.survey_button = QtWidgets.QPushButton(self)
		self.survey_button.setGeometry(QtCore.QRect(30, 110, 181, 41))
		self.survey_button.setObjectName("survey_button")
		self.survey_selected_label = QtWidgets.QLabel(self)
		self.survey_selected_label.setGeometry(QtCore.QRect(40, 160, 300, 16))
		self.survey_selected_label.setObjectName("survey_selected_label")
		self.select_survey_label = QtWidgets.QLabel(self)
		self.select_survey_label.setGeometry(QtCore.QRect(30, 80, 141, 21))
		font = QtGui.QFont()
		font.setBold(True)
		self.select_survey_label.setFont(font)
		self.select_survey_label.setObjectName("select_survey_label")
		self.survey_sheet_box = QtWidgets.QComboBox(self)
		self.survey_sheet_box.setGeometry(QtCore.QRect(290, 110, 161, 41))
		self.survey_sheet_box.setObjectName("survey_sheet_box")
		self.survey_sheet_label = QtWidgets.QLabel(self)
		self.survey_sheet_label.setGeometry(QtCore.QRect(290, 80, 141, 21))
		self.survey_sheet_label.setObjectName("survey_sheet_label")
		self.master_selected_label = QtWidgets.QLabel(self)
		self.master_selected_label.setGeometry(QtCore.QRect(40, 280, 300, 16))
		self.master_selected_label.setObjectName("master_selected_label")
		self.master_sheet_box = QtWidgets.QComboBox(self)
		self.master_sheet_box.setGeometry(QtCore.QRect(290, 230, 161, 41))
		self.master_sheet_box.setObjectName("master_sheet_box")
		self.select_master_label = QtWidgets.QLabel(self)
		self.select_master_label.setGeometry(QtCore.QRect(30, 200, 141, 21))
		font = QtGui.QFont()
		font.setBold(True)
		self.select_master_label.setFont(font)
		self.select_master_label.setObjectName("select_master_label")
		self.master_button = QtWidgets.QPushButton(self)
		self.master_button.setGeometry(QtCore.QRect(30, 230, 181, 41))
		self.master_button.setObjectName("master_button")
		self.master_sheet_label = QtWidgets.QLabel(self)
		self.master_sheet_label.setGeometry(QtCore.QRect(290, 200, 141, 21))
		self.master_sheet_label.setObjectName("master_sheet_label")
		self.back_button = QtWidgets.QPushButton(self)
		self.back_button.setGeometry(QtCore.QRect(5, 320, 80, 41))
		self.back_button.setObjectName("master_button")

		self.retranslateUi()
		QtCore.QMetaObject.connectSlotsByName(self)

	def retranslateUi(self):
		_translate = QtCore.QCoreApplication.translate
		self.setWindowTitle(_translate("MainWindow", "MainWindow"))
		self.title_label.setText(_translate("MainWindow", "Please select the Survey Sheet and the Master Sheet"))
		self.survey_button.setText(_translate("MainWindow", "Open File"))
		self.survey_selected_label.setText(_translate("MainWindow", "Selected File:"))
		self.select_survey_label.setText(_translate("MainWindow", "Select the Survey File"))
		self.survey_sheet_label.setText(_translate("MainWindow", "Sheet Name"))
		self.master_selected_label.setText(_translate("MainWindow", "Selected File:"))
		self.select_master_label.setText(_translate("MainWindow", "Select the Master File"))
		self.master_button.setText(_translate("MainWindow", "Open File"))
		self.back_button.setText(_translate("MainWindow", "Back"))
		self.master_sheet_label.setText(_translate("MainWindow", "Sheet Name"))

		self.survey_button.clicked.connect(lambda: self.upload_file(self.survey_selected_label, self.survey_sheet_box))
		self.master_button.clicked.connect(lambda: self.upload_file(self.master_selected_label, self.master_sheet_box))
	
	def upload_file(self, selected_label, combo_box):
		file_path = self.__file_explorer()
		if file_path:
			selected_label.setText(f"Selected File: {file_path}")
			workbook = load_workbook(file_path)
			sheet_names = workbook.sheetnames
			combo_box.addItems(sheet_names)

	def __file_explorer(self):
		fname=QFileDialog.getOpenFileName(self, 'Open file', 'C:\\', 'Excel File (*.xlsx *.xls)')
		return fname[0]


		
class Ui_MainWindow(QtWidgets.QMainWindow):
	def __init__(self):
		super().__init__()
		self.stacked_widget = QtWidgets.QStackedWidget()
		self.setCentralWidget(self.stacked_widget)
		
		self.folder_setter = Set_Folder()
		self.upload_file = Upload_Files()
		self.stacked_widget.addWidget(self.folder_setter)
		self.stacked_widget.addWidget(self.upload_file)
		self.Init_UI()

	def Init_UI(self):
		self.resize(592, 368)
		self.stacked_widget.setCurrentWidget(self.folder_setter)
		self.retranslateUI()
		QtCore.QMetaObject.connectSlotsByName(self)

	def retranslateUI(self):
		self.folder_setter.continue_button.clicked.connect(self.navigate_to_upload)
		self.upload_file.back_button.clicked.connect(lambda : self.navigate(self.folder_setter))
	
	def navigate_to_upload(self):
		try:
			with open('config.json', 'r') as file:
				config_dict = json.load(file)
				assert os.path.exists(config_dict['output_folder']), "Output folder does not exist"
			self.stacked_widget.setCurrentWidget(self.upload_file)
		except:
			_translate = QtCore.QCoreApplication.translate
			self.folder_setter.title_label.setText(_translate("MainWindow", "You have not set your storage folder, please select the folder to store the files"))

	def navigate(self, page, can_navigate=True):
		if can_navigate:
			self.stacked_widget.setCurrentWidget(page)
		
if __name__ == "__main__":
	app = QtWidgets.QApplication(sys.argv)
	home = Ui_MainWindow()
	home.show()
	sys.exit(app.exec_())
